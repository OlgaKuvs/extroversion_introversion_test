from openpyxl import load_workbook
from colorama import just_fix_windows_console
from colorama import Fore, Back, Style
from colorama import init
import os
import platform
import re


EXCEL_SHEET_NAME = 'test_e_i.xlsx'
BANNER = """
.---..   ..---..--.  .--.   --.--.   ..---..--.  .--.   .---..---..-..---.
|     \\ /   |  |   ):    :    |  |\\  |  |  |   ):    :    |  |   (   ) |  
|---   /    |  |--' |    |    |  | \\ |  |  |--' |    |    |  |--- `-.  |  
|     / \\   |  |  \ :    ;    |  |  \\|  |  |  \\ :    ;    |  |   (   ) |  
'---''   '  '  '   ` `--'   --'--'   '  '  '   ` `--'     '  '---'`-'  '  
'||''''|'||' '||''||'''||''|. ..|''||    '||'|.   '||''||'''||''|. ..|''||    |''||'''||''''| .|'''.|''||''| 
 ||  .    || |    ||   ||   |.|'    ||    || |'|   |   ||   ||   |.|'    ||      ||   ||  .   ||..  '  ||    
 ||''|     ||     ||   ||''|'||      ||   || | '|. |   ||   ||''|'||      ||     ||   ||''|    ''|||.  ||    
 ||       | ||    ||   ||   |'|.     ||   || |   |||   ||   ||   |'|.     ||     ||   ||     .     '|| ||    
.||......|   ||. .||. .||.  '|''|...|'   .||.|.   '|  .||. .||.  '|''|...|'     .||. .||.....|'....|' .||.   
                                                                                                             
                                                                                                             """                                                                                                                                                 
                                                                                                                                                 
THANK_YOU = """ 
.---..   .    .    .   ..   .  .   ..--. .   ..
  |  |   |   / \   |\\  ||  /    \ /:    :|   ||
  |  |---|  /- -\  | \\ ||-'      : |    ||   ||
  |  |   | /     \ |  \\||  \     | :    ;:   ;'
  '  '   ''       `'   ''   `    '  `--'  `-' o
"""

just_fix_windows_console()
init(autoreset=True)
regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b'
cursor_shape = '\x1b[3 q'
print(cursor_shape, end='')


def clear():
    # Check os and clear the terminal
    plt = platform.system()

    if plt == "Windows":
        def clear():
            return os.system('cls')
    elif plt == "Linux":
        def clear():
            return os.system('clear')
    clear()

# Class for the questions and answer key instances and used question flag


class Questions_Answers:
    def __init__(self):
        self.questions = ""
        self.answers = ""
        self.question_used = 0


print(Fore.YELLOW + Style.BRIGHT + BANNER +
      " WELCOME TO EXTROVERSION INTROVERSION TEST! \n ")


def check_data_workbook(name_used, email_used):
    """
    Check if user's data is already in the worksheet.
    If yes, print user's last test result.
    """
    try:
        wb2 = load_workbook(EXCEL_SHEET_NAME)
    except FileNotFoundError as fnf_error:
        print(Fore.RED + Style.BRIGHT + fnf_error)
    else:
        ws = wb2["Users"]

        for row in ws:
            name = str(row[0].value)
            email = row[1].value
            result = row[2].value
            if name.lower() == name_used.lower() and email == email_used:
                if result > -3 and result < 3:
                    print(f" Welcome again {name_used}!" +
                          "Your last result was mostly AMBIVERT")
                elif result <= -3:
                    print(f" Welcome again {name_used}!" +
                          " Your last result was mostly INTROVERT")
                elif result >= 3:
                    print(f" Welcome again {name_used}!" +
                          " Your last result was mostly EXTROVERT")
                return True


def check_data():

    """
    Check input for user's name and e-mail validation.
    If valid, call start_test function.
    """

    while True:
        name = input("\n Please enter your name:\n")
        name_valid = validate_name(name)

        if name_valid:
            print(Fore.GREEN + Style.BRIGHT +
                  f"\n Hello, {name}!\n")
            while True:
                email_str = input(" Please enter your email: \n")
                if validate_email(email_str):
                    while True:
                        clear()
                        start_test(name, email_str)
                else:
                    print(Fore.RED + Style.BRIGHT +
                          f" ❌ Invalid e-mail {email_str}..." +
                          " Please enter correct e-mail\n")
        else:
            print(Fore.RED + Style.BRIGHT +
                  f" ❌ Invalid name {name}... Please enter correct name\n")


def validate_name(name_val):
    """
    Splits a string (name) into a list and then join
    the list back into one string without spaces.
    Check if all the characters are alphabet letters.
    """

    name_check = "".join(name_val.split())
    if name_check.isalpha() and len(name_check) > 1:
        return True
    else:
        return False


def validate_email(email_val):
    """ 
    Check if email address is valid or not.       
    """
    x = re.fullmatch(regex, email_val)
    if x:
        return True    
        
    else:     
        return False   
      
 

def start_test(name_tested, email_tested):
    """
    Starting test, loading lists of questions 
    """    
    user_exists = check_data_workbook(name_tested, email_tested)  

    while True: 
         
        result = finish_test(user_exists)

        if result:   
            print(" - - - - - - - - -  - - - - - "+
            "- - - - - - - - - - - - - - - - - - - - - - ")
            print(Fore.GREEN + f" Welcome, {name_tested} 🙂 !" + 
            " Let's start. 🚀 \n Are you oriented more towards"+
            " the outer world or the inner world? 🤔\n ")    
            print(Fore.GREEN + " This easy test can give you a clear answer " + 
            "and help you understand \n your personality.🧑\n ")
            print(Fore.GREEN + Style.BRIGHT + " Please enter Y for 'YES'"+
            " answer or N for 'NO' answer. Enter Q to Quit")
            print(" - - - - - - - - -  - - - - - - - - -"+
            " - - - - - - - - - - - - - - - - - - \n")    

            list_t_normal = load_from_workbook("Test1")
            list_t_intra = load_from_workbook("Test2")
            list_t_extra = load_from_workbook("Test3")
            
            final_score = check_answers(list_t_normal, list_t_intra, list_t_extra)

            check_results(final_score)

            insert_user_data(name_tested, email_tested, final_score)
            user_exists = True      
                
        else:
            print(Fore.YELLOW + Style.BRIGHT + THANK_YOU)
            exit()


def insert_user_data(name_in, email_in, result_in):
    # Open worksheet
    try:
        wb2 = load_workbook(EXCEL_SHEET_NAME)                     
    except FileNotFoundError as fnf_error:
        print(Fore.RED  + fnf_error)
    else:
        ws = wb2["Users"] 

    # Find if user exists and overwrite test result
    row = ws.max_row + 1 
    user_found = 0    
    for i in range(1,row):
        s = str(ws.cell(i,1).value)
        e = str(ws.cell(i,2).value)
        if s.lower() == name_in.lower() and e == email_in:
            ws.cell(i,3).value = result_in
            user_found = row
    # If new user insert data to the find first empty row
    if user_found != row:
            ws.cell(row=row, column=1).value = name_in
            ws.cell(row=row, column=2).value = email_in
            ws.cell(row=row, column=3).value = result_in                  
    wb2.save(EXCEL_SHEET_NAME)      

def get_next_question(inner_score, list_test_normal,
    list_test_intra, list_test_extra):    
    #Сhoose a category of questions depending on the score.
        
    if inner_score > -3 and inner_score < 3:
        for i in range(len(list_test_normal)):
            if list_test_normal[i].question_used == 0:
                return list_test_normal[i]
    elif inner_score <= -3:
        for i in range(len(list_test_intra)):
            if list_test_intra[i].question_used == 0:
                return list_test_intra[i]
    elif inner_score >= 3:
        for i in range(len(list_test_extra)):
            if list_test_extra[i].question_used == 0:
                return list_test_extra[i]
    return None


def check_answers(list_test_normal, list_test_intra, list_test_extra): 
    """
    Run a while loop to check user's answers, 
    if 'Yes' or 'No', continue test, if 'Q' - quit.
    The loop will repeatedly print questions, 
    until the end of questions list.
    Check invalid input (any character except 'Y', 'N' or 'Q') 
    and return the error.
    """   
    score = 0  
   
    while True:         
            test_item = get_next_question(score, list_test_normal,
            list_test_intra, list_test_extra)            

            if test_item is None:                
                return score
            key_answer = test_item.answers 

            print(Fore.YELLOW + 
            f" {test_item.questions}  Enter  Y / N , Q to Quit")             
            user_answer = input("🔻\n")                               
           
            if user_answer.lower() == "q":
                print(Fore.YELLOW + Style.BRIGHT + THANK_YOU)                
                exit() 
            elif (user_answer.lower() == "y" and key_answer == 1) :       
                score += 1               
                test_item.question_used = 1      
            elif (user_answer.lower() == "y" and key_answer == 0):
                score -= 1               
                test_item.question_used = 1               
            elif (user_answer.lower() == "n" and key_answer == 1) :          
                score -= 1                
                test_item.question_used = 1                
            elif (user_answer.lower() == "n" and key_answer == 0):
                score += 1               
                test_item.question_used = 1               
            else:
                print(Fore.RED + "❌ Invalid data... " + 
                " Please enter  Y / N or Q to Quit \n")          
    
def check_results(score):     
    
    #Check results depending on score and print them.
    
    clear()

    print(" - - - - - - - - -  - - - - - - - - "+
          "- - - - - - - - - - - - - - - - - - - ")
    if score >= -3 and score <= 3:
        
        print(Fore.GREEN + Style.BRIGHT + 
        "\n Congratulations! You finished the test.\n\n You are mostly AMBIVERT,"+
        " \n exhibit qualities of both introversion and extroversion,"+
        "\n you can flip into either depending on their mood, context and goals.") 
        
    elif score < -3:
        print(Fore.GREEN + Style.BRIGHT + 
        "\n Congratulations! You finished the test.\n\n You are mostly INTROVERT,"+
        " \n you enjoy spending time alone and you feel more comfortable"+
        " \n focusing on your inner thoughts and ideas. ") 
        
    elif score > 3:
        print(Fore.GREEN + Style.BRIGHT + 
        "\n Congratulations! You finished the test.\n\n"+
        " You are mostly EXTROVERT, \n you enjoy being around "+
        "other people and you gain energy from them.")
    print(" - - - - - - - - -  - - - - - - - - - - - - - - - - - - - - - - - - - - - ")

def finish_test(user_is):     
    """
      Ask the user if they want to try the test 
      again and get the result (yes or no)
    """    

    if not user_is:        
        return True
    else:
        print(Fore.YELLOW + 
        "\n Would you like to try the test again?")

        while True:
            repeat_test = input(" Enter Y or N\n")

            if repeat_test.lower() == "y":
                clear()
                return True            
            elif repeat_test.lower() == "n":
                clear()
                return False  
            else:
                print(Fore.RED + 
                " ❌ Invalid data... Please enter Y or N \n")           
        
  
def load_from_workbook(test_sheet):
    # open worksheet
    try:
        wb2 = load_workbook(EXCEL_SHEET_NAME)                     
    except FileNotFoundError as fnf_error:
        print(Fore.WHITE + Back.RED + fnf_error)
    else:
        ws = wb2[test_sheet]

    list = [] 
    row = ws.max_row + 1 

    # populate list with object which contains questions, answer keys and flag     
    for i in range(1,row):
        qa = Questions_Answers()
        qa.questions = ws.cell(i,1).value
        qa.answers = ws.cell(i,2).value
        qa.question_used = 0       
        list.append(qa)
    list.pop(0)     
    return list                                

check_data() 

 




